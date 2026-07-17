import os
from datetime import datetime
from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, KeepTogether, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

class NumberedCanvas(canvas.Canvas):
    """
    Lienzo personalizado para ReportLab que realiza dos pasadas
    para calcular el total de páginas y pintar cabeceras y pies elegantes.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        width, height = self._pagesize
        
        # Obtener información dinámica de la empresa
        from apps.empresa.models import Empresa
        empresa = Empresa.objects.first()
        nombre_empresa = empresa.nombre.upper() if (empresa and empresa.nombre) else "SISMEING"
        
        # --- Cabecera ---
        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(colors.HexColor('#2B2B2B'))
        self.drawString(36, height - 30, nombre_empresa)
        
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor('#6B7280'))
        local_time = timezone.localtime(timezone.now())
        self.drawRightString(width - 36, height - 30, f"Generado: {local_time.strftime('%d/%m/%Y %H:%M')}")
        
        self.setStrokeColor(colors.HexColor('#E4E9F2'))
        self.setLineWidth(0.75)
        self.line(36, height - 35, width - 36, height - 35)
        
        # --- Pie de página ---
        self.line(36, 45, width - 36, 45)
        if empresa:
            footer_text = f"{empresa.nombre} - NIT: {empresa.nit} - Dir: {empresa.direccion} - Telf: {empresa.telefono}"
        else:
            footer_text = "Este documento es un reporte del sistema."
        self.drawString(36, 30, footer_text)
        
        page_text = f"Página {self._pageNumber} de {page_count}"
        self.drawRightString(width - 36, 30, page_text)
        
        self.restoreState()


def generar_pdf_reporte(titulo, headers, data, filename):
    """
    Genera un archivo PDF estructurado con ReportLab y lo devuelve en un HttpResponse.
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=54
    )

    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    style_title = ParagraphStyle(
        name='TitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#B22222'),
        spaceAfter=15,
        alignment=0 # Izquierda
    )
    
    style_header = ParagraphStyle(
        name='TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=0
    )
    
    style_cell = ParagraphStyle(
        name='TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#2B2B2B'),
        leading=10
    )

    elements = []

    # Logo de la empresa si existe
    from apps.empresa.models import Empresa
    empresa = Empresa.objects.first()
    
    logo_flowed = False
    if empresa and empresa.logo and os.path.exists(empresa.logo.path):
        try:
            img = Image(empresa.logo.path, width=120, height=35)
            img.hAlign = 'LEFT'
            elements.append(img)
            elements.append(Spacer(1, 10))
            logo_flowed = True
        except Exception:
            pass
            
    if not logo_flowed:
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo.png')
        if os.path.exists(logo_path):
            try:
                img = Image(logo_path, width=120, height=35)
                img.hAlign = 'LEFT'
                elements.append(img)
                elements.append(Spacer(1, 10))
            except Exception:
                pass

    # Título y metadatos
    elements.append(Paragraph(titulo, style_title))
    elements.append(Spacer(1, 10))

    # Construir tabla
    table_data = []
    
    # Cabeceras
    header_row = [Paragraph(h, style_header) for h in headers]
    table_data.append(header_row)
    
    # Datos
    for row in data:
        data_row = []
        for cell in row:
            text = str(cell) if cell is not None else ""
            data_row.append(Paragraph(text, style_cell))
        table_data.append(data_row)

    # Ancho dinámico para las columnas del PDF
    col_count = len(headers)
    page_width = letter[0] - 72 # 540 ptos disponibles
    col_width = page_width / col_count
    
    t = Table(table_data, colWidths=[col_width] * col_count)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#B22222')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E4E9F2')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    
    elements.append(t)
    
    # Construir el documento usando NumberedCanvas
    doc.build(elements, canvasmaker=NumberedCanvas)
    return response


def generar_excel_reporte(titulo, headers, data, filename):
    """
    Genera un archivo Excel formateado profesionalmente con openpyxl.
    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Mostrar líneas de cuadrícula
    ws.views.sheetView[0].showGridLines = True

    # Estilos de celdas
    font_title = Font(name="Calibri", size=16, bold=True, color="B22222")
    font_meta = Font(name="Calibri", size=10, italic=True, color="555555")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11, color="000000")
    
    fill_header = PatternFill(start_color="B22222", end_color="B22222", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='E4E9F2'),
        right=Side(style='thin', color='E4E9F2'),
        top=Side(style='thin', color='E4E9F2'),
        bottom=Side(style='thin', color='E4E9F2')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Escribir Título
    ws['A1'] = titulo
    ws['A1'].font = font_title
    
    # Escribir Metadatos
    from apps.empresa.models import Empresa
    empresa = Empresa.objects.first()
    nombre_empresa = empresa.nombre if (empresa and empresa.nombre) else "SISMEING"
    fecha_gen = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")
    ws['A2'] = f"Reporte de {nombre_empresa} generado automáticamente el {fecha_gen}"
    ws['A2'].font = font_meta
    
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 10 # Fila vacía de separación

    # Escribir Cabeceras
    row_idx = 4
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin
    
    ws.row_dimensions[row_idx].height = 24

    # Escribir Datos
    for r_data in data:
        row_idx += 1
        ws.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(r_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_thin
            
            # Cebra
            if row_idx % 2 == 0:
                cell.fill = fill_zebra
            
            # Formatear y alinear
            if isinstance(val, (int, float)):
                cell.alignment = align_right
                if isinstance(val, float):
                    cell.number_format = '#,##0.00'
            elif isinstance(val, (datetime, timezone.datetime)):
                cell.alignment = align_center
                cell.value = val.strftime('%d/%m/%Y %H:%M')
            else:
                cell.alignment = align_left

    # Autoajustar columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Saltar la primera fila porque contiene el título largo
            if cell.row in [1, 2]:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(response)
    return response


def generar_pdf_recibo(venta, detalles):
    """
    Genera el comprobante de venta formal en PDF con reportlab.
    Diseñado para hojas tamaño carta de manera elegante.
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="recibo_venta_{venta.id_venta}.pdf"'

    # Margen un poco más ajustado para factura
    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        leftMargin=45,
        rightMargin=45,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()
    
    # Modificar estilos existentes o crear nuevos
    style_company_name = ParagraphStyle(
        name='CompanyName',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#B22222'),
        leading=24
    )
    
    style_receipt_title = ParagraphStyle(
        name='ReceiptTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#2B2B2B'),
        alignment=2 # Derecha
    )
    
    style_label = ParagraphStyle(
        name='LabelStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.HexColor('#555555')
    )
    
    style_value = ParagraphStyle(
        name='ValueStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#111111')
    )
    
    style_th = ParagraphStyle(
        name='ThStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=0
    )
    
    style_td = ParagraphStyle(
        name='TdStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#2B2B2B'),
        leading=11
    )
    
    style_total_lbl = ParagraphStyle(
        name='TotalLblStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor('#111111'),
        alignment=2
    )

    style_total_val = ParagraphStyle(
        name='TotalValStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor('#B22222'),
        alignment=2
    )

    elements = []

    # --- ENCABEZADO DE LA FACTURA ---
    # Fila superior: Nombre/Logo a la izquierda, Título/Nro a la derecha
    from apps.empresa.models import Empresa
    empresa = Empresa.objects.first()
    
    logo_flowable = ""
    if empresa and empresa.logo and os.path.exists(empresa.logo.path):
        try:
            logo_flowable = Image(empresa.logo.path, width=120, height=35)
            logo_flowable.hAlign = 'LEFT'
        except Exception:
            pass
    else:
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo.png')
        if os.path.exists(logo_path):
            try:
                logo_flowable = Image(logo_path, width=120, height=35)
                logo_flowable.hAlign = 'LEFT'
            except Exception:
                pass

    nombre_empresa = empresa.nombre if (empresa and empresa.nombre) else "SISMEING"
    nit_str = f"NIT: {empresa.nit}" if (empresa and empresa.nit) else "NIT: S/N"
    direccion_str = f"Dirección: {empresa.direccion}" if (empresa and empresa.direccion) else "Dirección: Zona Central, Calle Principal #123"
    
    if empresa and empresa.telefono:
        telefono_str = f"Teléfono: {empresa.telefono} - {empresa.ciudad}"
    else:
        telefono_str = "Teléfono: (+591) 76543210 - Cochabamba, Bolivia"

    left_flow = [
        logo_flowable or Paragraph(nombre_empresa, style_company_name),
        Paragraph("Ensambles, Componentes y Computadoras", ParagraphStyle('SubC', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#777777'))),
        Paragraph(f"{direccion_str} | {nit_str}", ParagraphStyle('SubC2', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#777777'))),
        Paragraph(telefono_str, ParagraphStyle('SubC3', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#777777'))),
    ]
    
    right_flow = [
        Paragraph("COMPROBANTE DE VENTA", style_receipt_title),
        Spacer(1, 4),
        Paragraph(f"<b>Nro. Venta:</b> {str(venta.id_venta)[:18].upper()}...", ParagraphStyle('VId', parent=styles['Normal'], alignment=2, fontSize=9)),
        Paragraph(f"<b>Fecha:</b> {venta.fecha.strftime('%d/%m/%Y')}", ParagraphStyle('VFecha', parent=styles['Normal'], alignment=2, fontSize=9)),
    ]

    header_table = Table([[left_flow, right_flow]], colWidths=[270, 252])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    
    elements.append(header_table)
    
    # Línea divisoria decorativa
    dec_table = Table([[""]], colWidths=[522], rowHeights=[3])
    dec_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#B22222')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(dec_table)
    elements.append(Spacer(1, 15))

    # --- DATOS DEL CLIENTE ---
    cli = venta.cliente
    usr = venta.usuario
    
    cliente_info_data = [
        [Paragraph("<b>Cliente:</b>", style_label), Paragraph(f"{cli.nombre} {cli.apellido}", style_value),
         Paragraph("<b>Atendido por:</b>", style_label), Paragraph(f"{usr.nombre} {usr.apellido}", style_value)],
        [Paragraph("<b>NIT/CI:</b>", style_label), Paragraph(cli.ci or "S/N", style_value),
         Paragraph("<b>Fecha Registro:</b>", style_label), Paragraph(venta.created_at.strftime("%d/%m/%Y %H:%M"), style_value)]
    ]
    
    cliente_table = Table(cliente_info_data, colWidths=[70, 191, 90, 171])
    cliente_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#E4E9F2')),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E4E9F2')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    elements.append(cliente_table)
    elements.append(Spacer(1, 20))

    # --- DETALLE DE LA VENTA ---
    table_items_data = []
    # Cabecera tabla
    table_items_data.append([
        Paragraph("Cant.", style_th),
        Paragraph("Descripción del Producto", style_th),
        Paragraph("Precio Unitario", style_th),
        Paragraph("Descuento", style_th),
        Paragraph("Subtotal", style_th)
    ])
    
    for d in detalles:
        table_items_data.append([
            Paragraph(str(d.cantidad), style_td),
            Paragraph(d.producto.nombre, style_td),
            Paragraph(f"Bs. {d.precio_unitario:.2f}", style_td),
            Paragraph(f"Bs. {d.descuento:.2f}", style_td),
            Paragraph(f"Bs. {d.subtotal:.2f}", style_td)
        ])

    # Totales
    table_items_data.append([
        "", "", "",
        Paragraph("TOTAL:", style_total_lbl),
        Paragraph(f"Bs. {venta.total:.2f}", style_total_val)
    ])
    
    items_table = Table(table_items_data, colWidths=[50, 202, 90, 80, 100])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E1E1E')), # Cabecera oscura premium
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -2), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -2), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#E4E9F2')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8FAFC')]),
        # Estilo para la fila del total (SPAN de col 0 a 2)
        ('SPAN', (0, -1), (2, -1)),
        ('TOPPADDING', (3, -1), (-1, -1), 15),
        ('BOTTOMPADDING', (3, -1), (-1, -1), 15),
    ]))
    
    elements.append(items_table)
    elements.append(Spacer(1, 40))

    # --- PIE DE FACTURA ---
    pie_flow = [
        Paragraph("¡Gracias por su preferencia!", ParagraphStyle('Gracias', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, alignment=1, textColor=colors.HexColor('#B22222'))),
        Spacer(1, 4),
        Paragraph("Cualquier reclamo se realiza presentando este comprobante.", ParagraphStyle('Term', parent=styles['Normal'], fontSize=8, alignment=1, textColor=colors.HexColor('#6B7280'))),
        Paragraph(f"{nombre_empresa} - Soluciones Informáticas", ParagraphStyle('SysN', parent=styles['Normal'], fontSize=8, alignment=1, textColor=colors.HexColor('#9CA3AF'))),
    ]
    elements.append(KeepTogether(pie_flow))

    doc.build(elements)
    return response
